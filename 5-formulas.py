from openpyxl import load_workbook
from openpyxl.utils import get_column_letter #pegar uma letra da coluna

# 1- lê pasta de trabalho e planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatorio"] #indica qual planilha quero utilizar, no caso Relatorio 

# 2- referências das linhas e colunas (coluna min,max e linha min,max) || python vai verificar a planilha e trazer a informação
min_column = wb.active.min_column 
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3- incluindo formulas
for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency" #trabalhar com o valor da moeda

wb.save("test.xlsx")