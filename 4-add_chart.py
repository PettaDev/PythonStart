from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 1- lê pasta de trabalho e planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatorio"] #indica qual planilha quero utilizar, no caso Relatorio 

# 2- referências das linhas e colunas (coluna min,max e linha min,max) || python vai verificar a planilha e trazer a informação
#colunas
min_column = wb.active.min_column 
max_column = wb.active.max_column
#linhas
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3- adicionando dados e categorias no gráfico
barchart = BarChart() #instancialo

data = Reference(
   sheet,
   min_col=min_column + 1,
   max_col=max_column,
   min_row=min_row,
   max_row=max_row 
)

categories = Reference(
   sheet,
   min_col=min_column,
   max_col=min_column, #min_column para uma melhor interpretação do gráfico
   min_row=min_row + 1,
   max_row=max_row 
)


barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# 4- criando gráfico
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricante"
barchart.style = 3

# 5- salvando o workbook
wb.save("data/barchart.xlsx")